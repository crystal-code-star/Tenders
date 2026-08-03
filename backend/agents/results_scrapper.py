"""
marches_publics_scraper.py - Scraper Résultats AO Maroc
======================================================
Scrape les résultats définitifs des appels d'offres du portail marocain.
Les fichiers joints sont stockés EN BASE64 dans le fichier Excel.
Sauvegarde progressive dans Excel.
"""

import re
import time
import logging
import warnings
from datetime import datetime
from typing import Dict, Optional
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
import base64
import io

from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

warnings.filterwarnings("ignore")
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ─── CONFIGURATION ─────────────────────────────────────────
BASE_URL = "https://www.marchespublics.gov.ma"
SEARCH_URL = f"{BASE_URL}/index.php?page=entreprise.EntrepriseAdvancedSearch&AvisAttribution"
START_DATE = "01/01/2024"
END_DATE = datetime.now().strftime("%d/%m/%Y")
MAX_PAGES = None  # None = toutes les pages, ou mettre un nombre
OUTPUT_FILE = f"ao_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# ─── LOGGING ────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(f'scraper_{datetime.now().strftime("%Y%m%d")}.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)


class MarchesPublicsScraper:
    """Scraper pour les résultats d'appels d'offres du portail marocain."""
    
    def __init__(self):
        self.results_count = 0
        self.total_pages = 0
        self.current_page = 1
        self.output_file = OUTPUT_FILE
        
    def _init_excel(self):
        """Initialise le fichier Excel avec les en-têtes."""
        if os.path.exists(self.output_file):
            os.remove(self.output_file)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Résultats AO"
        
        # En-têtes
        headers = [
            'reference', 'procedure_abbr', 'procedure', 'categorie',
            'date_publication', 'objet', 'acheteur_public',
            'type_annonce', 'lieu_execution',
            'fichier_nom', 'fichier_base64', 'fichier_taille_kb', 'date_scraping'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
        
        # Ajuster la largeur des colonnes
        column_widths = [20, 15, 30, 15, 15, 50, 40, 20, 30, 30, 20, 15, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        wb.save(self.output_file)
        logger.info(f"📁 Fichier Excel initialisé: {self.output_file}")
        
    def _save_result_to_excel(self, result: Dict):
        """Sauvegarde immédiatement un résultat dans le fichier Excel."""
        try:
            wb = load_workbook(self.output_file)
            ws = wb.active
            next_row = ws.max_row + 1
            
            columns = [
                'reference', 'procedure_abbr', 'procedure', 'categorie',
                'date_publication', 'objet', 'acheteur_public',
                'type_annonce', 'lieu_execution',
                'fichier_nom', 'fichier_base64', 'fichier_taille_kb'
            ]
            
            for col, key in enumerate(columns, 1):
                value = result.get(key, '')
                
                # Pour la colonne base64, limiter si trop grand (Excel a une limite de cellule)
                if key == 'fichier_base64' and isinstance(value, str) and len(value) > 32000:
                    # Si trop grand, on tronque avec un message
                    value = value[:32000] + "...[TRONQUÉ - Fichier trop volumineux pour Excel]"
                    logger.warning(f"  ⚠️ Fichier base64 tronqué (>32000 caractères)")
                
                # Limiter la longueur pour éviter les problèmes Excel
                if isinstance(value, str) and len(value) > 32767:
                    value = value[:32767]
                
                ws.cell(row=next_row, column=col, value=value)
            
            # Date de scraping
            ws.cell(row=next_row, column=13, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            
            wb.save(self.output_file)
            
        except Exception as e:
            logger.error(f"❌ Erreur sauvegarde Excel: {e}")
    
    def _download_file_to_base64(self, page, detail_page, result: Dict) -> Dict:
        """
        Télécharge le fichier joint et le stocke en base64.
        Retourne un dict avec les infos du fichier.
        """
        fichier_info = {
            'fichier_nom': '',
            'fichier_base64': '',
            'fichier_taille_kb': ''
        }
        
        try:
            ref_consultation = result.get('ref_consultation', '')
            org_acronyme = result.get('org_acronyme', '')
            
            if not ref_consultation or not org_acronyme:
                logger.debug(f"  Pas d'infos pour télécharger le fichier")
                return fichier_info
            
            # 2. Naviguer vers la page de détail
            detail_full_url = f"{BASE_URL}/index.php?page=entreprise.EntrepriseDetailConsultation&refConsultation={ref_consultation}&orgAcronyme={org_acronyme}"
            
            logger.info(f"  🌐 Accès page détail: {ref_consultation}")
            
            detail_page.goto(detail_full_url, wait_until="domcontentloaded", timeout=30000)
            detail_page.wait_for_timeout(3000)
            
            # 3. Chercher le lien de téléchargement
            download_link = None
            download_text = ""
            
            # Chercher le lien avec l'icône picto-compresse
            selectors = [
                "li.picto-link a",
                "a[href*='EntrepriseDownloadAvis']",
                "a[href*='DownloadAvisJAL']",
                "a:has-text('Fichier joint')",
                "a:has-text('Télécharger')",
                "a[href*='download']",
            ]
            
            for selector in selectors:
                try:
                    elements = detail_page.query_selector_all(selector)
                    for elem in elements:
                        if elem.is_visible():
                            href = elem.get_attribute('href')
                            text = elem.inner_text().strip()
                            if href and ('DownloadAvis' in href or 'download' in href.lower() or 'Fichier' in text):
                                download_link = href
                                download_text = text
                                break
                    if download_link:
                        break
                except:
                    continue
            
            if not download_link:
                # Chercher l'image picto-compresse et son lien parent
                try:
                    img_compresse = detail_page.query_selector("img[src*='picto-compresse']")
                    if img_compresse:
                        # Remonter au li parent et chercher le lien
                        parent_link = detail_page.evaluate("""
                            (img) => {
                                const li = img.closest('li.picto-link');
                                if (li) {
                                    const a = li.querySelector('a');
                                    return a ? a.href : null;
                                }
                                return null;
                            }
                        """, img_compresse)
                        if parent_link:
                            download_link = parent_link
                            download_text = "Fichier joint"
                except:
                    pass
            
            if not download_link:
                logger.info(f"  ℹ️ Aucun fichier à télécharger pour ce résultat")
                return fichier_info
            
            # 4. Construire l'URL complète
            if download_link.startswith('?'):
                download_link = f"{BASE_URL}/index.php{download_link}"
            elif download_link.startswith('/'):
                download_link = f"{BASE_URL}{download_link}"
            elif not download_link.startswith('http'):
                download_link = f"{BASE_URL}/{download_link}"
            
            logger.info(f"  📥 Téléchargement: {download_text[:80]}")
            
            # 5. Télécharger le fichier en mémoire
            try:
                with detail_page.expect_download(timeout=60000) as download_info:
                    # Cliquer sur le lien
                    detail_page.click(f"a[href*='DownloadAvis']")
                
                download = download_info.value
                
                # Récupérer le contenu du fichier
                file_path = download.path()
                if not file_path:
                    logger.warning(f"  ⚠️ Chemin de fichier vide")
                    return fichier_info
                
                # Lire le fichier
                with open(file_path, 'rb') as f:
                    file_content = f.read()
                
                if not file_content:
                    logger.warning(f"  ⚠️ Fichier vide")
                    return fichier_info
                
                # Encoder en base64
                file_base64 = base64.b64encode(file_content).decode('utf-8')
                
                # Infos du fichier
                suggested_filename = download.suggested_filename or f"fichier_{ref_consultation}.pdf"
                file_size_kb = len(file_content) / 1024
                
                fichier_info = {
                    'fichier_nom': suggested_filename,
                    'fichier_base64': file_base64,
                    'fichier_taille_kb': f"{file_size_kb:.1f}"
                }
                
                logger.info(f"  ✅ Fichier encodé: {suggested_filename} ({file_size_kb:.1f} KB, base64: {len(file_base64)} chars)")
                
                # Avertir si le fichier est très gros
                if len(file_base64) > 1000000:  # > 1MB en base64
                    logger.warning(f"  ⚠️ Fichier volumineux! Base64: {len(file_base64) / 1024:.0f}KB - peut ralentir Excel")
                
                return fichier_info
                
            except PlaywrightTimeout:
                logger.warning(f"  ⚠️ Timeout téléchargement")
                return fichier_info
            except Exception as e:
                logger.error(f"  ❌ Erreur téléchargement: {e}")
                return fichier_info
                
        except Exception as e:
            logger.error(f"  ❌ Erreur accès page détail: {e}")
            return fichier_info
    
    def scrape(self):
        """Méthode principale de scraping."""
        logger.info("=" * 60)
        logger.info("🚀 Démarrage du scraping des résultats AO")
        logger.info(f"📅 Date début: {START_DATE}")
        logger.info(f"📅 Date fin: {END_DATE}")
        logger.info(f"💾 Sauvegarde (fichiers en base64): {self.output_file}")
        logger.info("=" * 60)
        
        # Initialiser le fichier Excel
        self._init_excel()
        
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=['--no-sandbox', '--disable-setuid-sandbox', '--disable-dev-shm-usage']
            )
            context = browser.new_context(
                viewport={"width": 1366, "height": 768},
                locale="fr-FR",
                timezone_id="Africa/Casablanca",
                accept_downloads=True
            )
            page = context.new_page()
            detail_page = context.new_page()
            
            try:
                # Navigation et recherche
                self._navigate_to_search(page)
                self._fill_search_form(page)
                self._submit_search(page)
                self._get_total_pages(page)
                
                # Scanner toutes les pages
                self._scan_all_pages(page, detail_page)
                
            except Exception as e:
                logger.error(f"❌ Erreur lors du scraping: {e}")
            finally:
                detail_page.close()
                browser.close()
        
        self._print_summary()
        
    def _navigate_to_search(self, page):
        """Navigue vers la page de recherche avancée."""
        logger.info("🌐 Navigation vers la page de recherche...")
        page.goto(SEARCH_URL, wait_until="networkidle", timeout=60000)
        page.wait_for_timeout(3000)
        logger.info("✅ Page chargée")
        
    def _fill_search_form(self, page):
        """Remplit le formulaire avec la date de début."""
        logger.info("📝 Remplissage du formulaire...")
        
        date_start_selector = "#ctl0_CONTENU_PAGE_AdvancedSearch_dateMiseEnLigneCalculeStart"
        try:
            page.fill(date_start_selector, "")
            page.fill(date_start_selector, START_DATE)
            logger.info(f"  Date début: {START_DATE}")
        except Exception as e:
            logger.warning(f"  ⚠️ Erreur remplissage date début: {e}")
        
        date_end_selector = "#ctl0_CONTENU_PAGE_AdvancedSearch_dateMiseEnLigneCalculeEnd"
        try:
            page.fill(date_end_selector, "")
            page.fill(date_end_selector, END_DATE)
            logger.info(f"  Date fin: {END_DATE}")
        except Exception as e:
            logger.warning(f"  ⚠️ Erreur remplissage date fin: {e}")
        
        page.wait_for_timeout(1000)
        
    def _submit_search(self, page):
        """Soumet le formulaire de recherche."""
        logger.info("🔍 Lancement de la recherche...")
        
        selectors = [
            "#ctl0_CONTENU_PAGE_AdvancedSearch_lancerRecherche",
            "input[value='Lancer la recherche']",
            "input[name*='lancerRecherche']"
        ]
        
        clicked = False
        for selector in selectors:
            try:
                page.click(selector, timeout=10000)
                clicked = True
                break
            except:
                continue
        
        if not clicked:
            page.evaluate("document.querySelector('form').submit()")
        
        page.wait_for_timeout(5000)
        try:
            page.wait_for_selector("table.table-results", timeout=30000)
            logger.info("✅ Résultats chargés")
        except PlaywrightTimeout:
            logger.warning("⚠️ Tableau des résultats non trouvé après timeout")
            
    def _get_total_pages(self, page):
        """Récupère le nombre total de pages."""
        try:
            nb_pages_elem = page.query_selector("#ctl0_CONTENU_PAGE_resultSearch_nombrePageTop")
            if nb_pages_elem:
                self.total_pages = int(nb_pages_elem.inner_text().strip())
                logger.info(f"📊 Nombre total de pages: {self.total_pages}")
            else:
                self.total_pages = 1
        except:
            self.total_pages = 1
            
        if MAX_PAGES:
            self.total_pages = min(self.total_pages, MAX_PAGES)
            logger.info(f"  Limité à {self.total_pages} pages")
            
    def _scan_all_pages(self, page, detail_page):
        """Parcourt toutes les pages de résultats."""
        logger.info(f"\n📄 Début du scan ({self.total_pages} pages)...\n")
        
        for page_num in range(1, self.total_pages + 1):
            self.current_page = page_num
            logger.info(f"\n{'='*50}")
            logger.info(f"📄 PAGE {page_num}/{self.total_pages}")
            logger.info(f"{'='*50}")
            
            count_before = self.results_count
            self._extract_page_results(page, detail_page)
            count_page = self.results_count - count_before
            
            logger.info(f"  ✅ {count_page} résultats traités (Total: {self.results_count})")
            
            if page_num < self.total_pages:
                self._go_to_next_page(page)
                
            time.sleep(2)
            
    def _extract_page_results(self, page, detail_page):
        """Extrait, télécharge (en base64) et sauvegarde les résultats."""
        html = page.content()
        soup = BeautifulSoup(html, 'html.parser')
        
        rows = soup.select("tr:has(td.col-450)")
        
        if not rows:
            logger.warning("  Aucune ligne trouvée sur cette page")
            return
        
        for i, row in enumerate(rows, 1):
            try:
                result = self._parse_row(row)
                
                if not result:
                    continue
                
                # Récupérer les infos pour le téléchargement
                actions_td = row.select_one("td.actions")
                if actions_td:
                    link = actions_td.select_one("a[href*='DetailConsultation']")
                    if link:
                        href = link.get("href", "")
                        ref_match = re.search(r'refConsultation=(\d+)', href)
                        org_match = re.search(r'orgAcronyme=([a-zA-Z0-9]+)', href)
                        
                        if ref_match:
                            result['ref_consultation'] = ref_match.group(1)
                        if org_match:
                            result['org_acronyme'] = org_match.group(1)
                
                # Télécharger le fichier en base64
                if result.get('ref_consultation') and result.get('org_acronyme'):
                    fichier_info = self._download_file_to_base64(page, detail_page, result)
                    result.update(fichier_info)
                
                # Sauvegarder immédiatement dans Excel
                self._save_result_to_excel(result)
                self.results_count += 1
                
                # Log
                ref = result.get('reference', 'N/A')
                objet = result.get('objet', '')[:60]
                cat = result.get('categorie', '')
                fichier_status = "📥" if result.get('fichier_base64') else "❌"
                
                logger.info(f"  [{self.results_count}] {fichier_status} {ref} | {cat} | {objet}...")
                
            except Exception as e:
                logger.debug(f"  Erreur traitement ligne {i}: {e}")
                continue
        
    def _parse_row(self, row) -> Optional[Dict]:
        """Parse une ligne du tableau de résultats."""
        result = {}
        
        # Référence
        ref_span = row.select_one("span.ref")
        if ref_span:
            result['reference'] = ref_span.get_text(strip=True)
        
        # Procédure (abréviation)
        proc_abbr_div = row.select_one("div.line-info-bulle")
        if proc_abbr_div:
            text = proc_abbr_div.get_text(strip=True).replace('...', '').strip()
            result['procedure_abbr'] = text
        
        # Procédure (nom complet)
        proc_full_div = row.select_one("div[id*='type_procedure']")
        if proc_full_div:
            result['procedure'] = proc_full_div.get_text(strip=True)
        
        # Catégorie
        cat_div = row.select_one("div[id*='panelBlocCategorie']")
        if cat_div:
            result['categorie'] = cat_div.get_text(strip=True)
        
        # Date de publication
        td_col90 = row.select_one("td.col-90")
        if td_col90:
            for div in td_col90.find_all("div"):
                text = div.get_text(strip=True)
                if re.match(r'\d{2}/\d{2}/\d{4}', text):
                    result['date_publication'] = text
                    break
        
        # Objet
        objet_div = row.select_one("div[id*='panelBlocObjet']")
        if objet_div:
            objet_text = objet_div.get_text(strip=True)
            objet_text = re.sub(r'^Objet\s*:\s*', '', objet_text).strip()
            result['objet'] = objet_text
        
        # Acheteur public
        acheteur_div = row.select_one("div[id*='panelBlocDenomination']")
        if acheteur_div:
            acheteur_text = acheteur_div.get_text(strip=True)
            acheteur_text = re.sub(r'^Acheteur\s+public\s*:\s*', '', acheteur_text).strip()
            result['acheteur_public'] = acheteur_text
        
        # Type d'annonce
        td_lieu = row.select("td.col-90")[1] if len(row.select("td.col-90")) > 1 else None
        if td_lieu:
            imgs = td_lieu.select("img")
            for img in imgs:
                src = img.get('src', '')
                title = img.get('title', '')
                if 'picto-avis-attribution' in src:
                    result['type_annonce'] = 'Résultat définitif'
                elif 'picto-PV' in src:
                    result['type_annonce'] = 'Extrait de PV'
                elif title:
                    result['type_annonce'] = title
            
            cloture_line = td_lieu.select_one("div.cloture-line")
            if cloture_line:
                detail_text = cloture_line.get_text(strip=True)
                if detail_text and not result.get('type_annonce'):
                    result['type_annonce'] = detail_text
        
        # Lieu d'exécution
        lieu_div = row.select_one("div[id*='panelBlocLieuxExec']")
        if lieu_div:
            lieu_text = lieu_div.get_text(separator=' ', strip=True)
            lieu_text = re.sub(r'\s+', ' ', lieu_text).strip()
            result['lieu_execution'] = lieu_text
        
        if not result.get('reference'):
            return None
        
        return result
    
    def _go_to_next_page(self, page):
        """Passe à la page suivante."""
        next_page = self.current_page + 1
        
        try:
            page_input = "#ctl0_CONTENU_PAGE_resultSearch_numPageTop"
            page.fill(page_input, "")
            page.fill(page_input, str(next_page))
            page.press(page_input, "Enter")
            page.wait_for_timeout(5000)
            page.wait_for_selector("table.table-results", timeout=30000)
        except Exception as e:
            logger.warning(f"  ⚠️ Erreur navigation page {next_page}: {e}")
            try:
                next_arrow = page.query_selector("a[id*='PagerTop_ctl2']")
                if next_arrow:
                    next_arrow.click()
                    page.wait_for_timeout(5000)
                    page.wait_for_selector("table.table-results", timeout=30000)
            except:
                logger.error(f"  ❌ Impossible d'aller à la page {next_page}")
                raise
    
    def _print_summary(self):
        """Affiche le résumé final."""
        logger.info("\n" + "=" * 60)
        logger.info("📊 RÉSUMÉ DU SCRAPING")
        logger.info("=" * 60)
        logger.info(f"✅ Fichier Excel: {self.output_file}")
        logger.info(f"📈 Total résultats: {self.results_count}")
        logger.info(f"📅 Période: {START_DATE} → {END_DATE}")
        logger.info(f"📄 Pages scannées: {self.current_page}/{self.total_pages}")
        
        if os.path.exists(self.output_file):
            size_mb = os.path.getsize(self.output_file) / (1024 * 1024)
            logger.info(f"💾 Taille du fichier Excel: {size_mb:.1f} MB")
        
        logger.info("=" * 60)
        logger.info("")
        logger.info("📋 Pour utiliser les fichiers depuis Excel (Python):")
        logger.info("   import base64")
        logger.info("   with open('fichier.pdf', 'wb') as f:")
        logger.info("       f.write(base64.b64decode(cell_value))")
        logger.info("")


def main():
    """Fonction principale."""
    scraper = MarchesPublicsScraper()
    scraper.scrape()


if __name__ == "__main__":
    main()
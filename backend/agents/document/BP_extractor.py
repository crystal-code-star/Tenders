"""
COMPLETE BP (Bordereau des Prix) EXTRACTION SYSTEM
==================================================
Version: 3.3 - Fixed + --force option
"""
import os, sys, re, json, logging, io, base64, zipfile, argparse, time
from pathlib import Path
from datetime import datetime, timezone
from typing import Dict, Any, List, Optional, Tuple, Union
from io import BytesIO
from dotenv import load_dotenv
from colorama import init, Fore, Style, Back

init(autoreset=True)
current_dir = Path(__file__).resolve().parent
env_path = current_dir.parent.parent / ".env"
load_dotenv(env_path)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(name)s: %(message)s')
logger = logging.getLogger("BP_extractor")

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")
if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
    print(f"\n{Fore.RED}❌ ERREUR: Identifiants Supabase manquants dans .env"); sys.exit(1)
import supabase
supabase_client = supabase.create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

def print_header(t): print(f"\n{Fore.CYAN}{Style.BRIGHT}{'═'*80}\n{Fore.CYAN}{Style.BRIGHT}  {t}\n{Fore.CYAN}{Style.BRIGHT}{'═'*80}")
def print_section(t): print(f"\n{Fore.YELLOW}{'─'*80}\n{Fore.YELLOW}  {t}\n{Fore.YELLOW}{'─'*80}")
def print_success(t): print(f"  {Fore.GREEN}✅ {t}")
def print_error(t): print(f"  {Fore.RED}❌ {t}")
def print_warning(t): print(f"  {Fore.YELLOW}⚠️  {t}")
def print_info(t): print(f"  {Fore.WHITE}📌 {t}")
def print_stat(t): print(f"  {Fore.MAGENTA}📊 {t}")

global_stats = {"total_tenders":0,"processed":0,"skipped_already_done":0,"errors":0,"total_items_extracted":0}

# ═══════════════ FILE DETECTION ═══════════════
def is_bp_file(filename: str) -> bool:
    filename_lower = filename.lower()
    bp_keywords = ['bp','bordereau','bpu','prix','bordereau des prix','dp','devis','estimatif','quantitatif','dq','bq','bordereau quantitatif','mercuriale','serie','prix unitaire','prix total']
    exclude_keywords = ['rc','acte','engagement','attestation','certificat','cv','curriculum','rib','bancaire','plan','planning','rapport','avis','aoo','ao','cahier','cps','ccp','cctp','ccag','reglement','consultation','prospectus','declaration','honneur','moyens','memoire','methodologie','echantillon','note']
    return any(kw in filename_lower for kw in bp_keywords) and not any(kw in filename_lower for kw in exclude_keywords)

# ═══════════════ NUMBER CLEANING ═══════════════
def _clean_number(value: str) -> Optional[float]:
    if not value or not str(value).strip(): return None
    value = str(value).strip(); value = re.sub(r'\s','',value)
    if ',' in value:
        parts = value.split(',')
        if len(parts) == 2: value = f"{parts[0].replace('.','')}.{parts[1]}"
        else: value = f"{''.join(parts[:-1]).replace('.','')}.{parts[-1]}"
    else: value = value.replace('.','')
    try: return float(value)
    except ValueError: return None

# ═══════════════ DISPLAY ═══════════════
def display_bp_items(items, max_display=20):
    if not items: print_warning("Aucun item"); return
    print_section(f"📋 ITEMS BP ({len(items)} total, {min(len(items),max_display)} affichés)")
    for i, item in enumerate(items[:max_display], 1):
        n_prix = item.get("N° Prix","")[:6]; designation = item.get("Désignation","")[:40]
        unite = item.get("Unité","")[:5]; quantite = item.get("Quantité","")[:8]
        pu_ht = item.get("Prix Unitaire HT","")[:10]; total_ht = item.get("Total HT","")[:10]
        print(f"  {Fore.CYAN}{i:<4} {Fore.WHITE}{n_prix:<6} {designation:<40} {unite:<5} {Fore.GREEN}{quantite:<8} {pu_ht:<10} {total_ht:<10}")

def display_bp_summary(doc_level, items_count):
    print_section("📊 RÉSUMÉ BP")
    for label, key in [("Réf. AO","Ref_AO"),("Objet","Objet"),("Maître d'ouvrage","Maitre_Ouvrage"),("Total HT","Total_HT")]:
        val = doc_level.get(key,'N/A')
        print(f"  {Fore.CYAN}{label:<20}: {Fore.WHITE}{str(val)[:50]}")

# ═══════════════ ZIP EXTRACTION ═══════════════
def extract_files_from_zip(zip_data) -> List[Dict[str, Any]]:
    try:
        zip_bytes = base64.b64decode(zip_data) if isinstance(zip_data, str) else zip_data
        with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
            all_files = zf.namelist()
            bp_files = [f for f in all_files if is_bp_file(Path(f).name)]
            if not bp_files:
                logger.info("Recherche Excel...")
                bp_files = [f for f in all_files if f.lower().endswith(('.xlsx','.xls','.xlsm'))]
            if not bp_files:
                logger.info("Recherche Word...")
                bp_files = [f for f in all_files if f.lower().endswith(('.docx','.doc'))]
            if not bp_files:
                logger.info("Recherche PDF...")
                bp_files = [f for f in all_files if f.lower().endswith('.pdf')]
            return [{"filename":Path(f).name,"full_path":f,"file_bytes":zf.read(f),"size_kb":len(zf.read(f))/1024} for f in bp_files[:10]]
    except Exception as e:
        logger.error(f"Erreur ZIP: {e}")
        return []

# ═══════════════ SUPABASE ═══════════════
def get_all_tenders_to_process(force=False):
    print_header("RECHERCHE AO" + (" (MODE FORCÉ)" if force else ""))
    try:
        r1 = supabase_client.table("tenders_3").select("reference,objet,acheteur_public,dce_zip_base64,dce_zip_url,bp_extraction_status").not_.is_("dce_zip_base64","null").execute()
        r2 = supabase_client.table("tenders_3").select("reference,objet,acheteur_public,dce_zip_base64,dce_zip_url,bp_extraction_status").not_.is_("dce_zip_url","null").execute()
        all_t = {}
        for t in (r1.data or []) + (r2.data or []):
            ref = t.get('reference')
            if ref and ref not in all_t: all_t[ref] = t
        to_process = []
        for ref, t in all_t.items():
            if force: to_process.append(t)
            elif t.get('bp_extraction_status') != 'completed': to_process.append(t)
            else: global_stats["skipped_already_done"] += 1
        global_stats["total_tenders"] = len(all_t)
        print_success(f"Total: {len(all_t)} | À traiter: {len(to_process)}")
        return to_process
    except Exception as e:
        print_error(f"Erreur: {e}"); return []

def get_tender_files(reference: str):
    try:
        r = supabase_client.table("tenders_3").select("*").eq("reference",reference).execute()
        if not r.data: return None
        t = r.data[0]
        b64 = t.get('dce_zip_base64'); url = t.get('dce_zip_url')
        if b64: files = extract_files_from_zip(b64)
        elif url:
            import requests as req
            try:
                resp = req.get(url, timeout=30, stream=True)
                if resp.status_code == 200:
                    content = b""
                    for chunk in resp.iter_content(chunk_size=8192):
                        content += chunk
                        if len(content) > 50*1024*1024: break
                    files = extract_files_from_zip(content)
                else: return None
            except Exception as e:
                logger.error(f"Téléchargement échoué: {e}")
                return None
        else: return None
        return {"tender":t,"files":files,"reference":reference}
    except Exception as e:
        logger.error(f"Erreur: {e}"); return None

def save_bp_to_supabase(reference, bp_result, filename):
    try:
        items = bp_result.get("items",[])
        supabase_client.table("tenders_3").update({"bp_extraction_status":"completed","bp_extracted_at":datetime.now(timezone.utc).isoformat()}).eq("reference",reference).execute()
        try: supabase_client.table("tenders_3_bp_items").delete().eq("tender_reference",reference).execute()
        except: pass
        if items:
            batch = []
            for item in items:
                n_prix = item.get("N° Prix","").strip(); designation = item.get("Désignation","").strip()
                if not n_prix and not designation: continue
                batch.append({"tender_reference":reference,"n_prix":n_prix or None,"designation":designation or None,"unite":item.get("Unité","").strip() or None,"quantite":_clean_number(item.get("Quantité","")),"prix_unitaire_ht":_clean_number(item.get("Prix Unitaire HT","")),"total_ht":_clean_number(item.get("Total HT",""))})
            for i in range(0,len(batch),100): supabase_client.table("tenders_3_bp_items").insert(batch[i:i+100]).execute()
            logger.info(f"✅ {len(batch)} items sauvegardés")
        return True
    except Exception as e:
        logger.error(f"Erreur sauvegarde: {e}")
        try: supabase_client.table("tenders_3").update({"bp_extraction_status":"error"}).eq("reference",reference).execute()
        except: pass
        return False

# ═══════════════ BP EXTRACTION SIMPLIFIÉE ═══════════════
def extract_bp_from_excel(file_bytes, ext=".xlsx"):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        all_items = []
        for sn in wb.sheetnames:
            ws = wb[sn]; grid = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells): grid.append(cells)
            if not grid: continue
            # Trouver header
            header_idx = None
            for idx, row in enumerate(grid[:25]):
                row_text = " ".join(str(c) for c in row if c).lower()
                if sum(1 for t in ["prix","n°","designation","unité","quantité","total"] if t in row_text) >= 3:
                    header_idx = idx; break
            if header_idx is None: continue
            header = grid[header_idx]
            # Mapping simple
            col_map = {}
            for i, cell in enumerate(header):
                ct = str(cell).lower().strip()
                if any(k in ct for k in ['n°','num','article','item','n ']): col_map[i] = "N° Prix"
                elif any(k in ct for k in ['designation','description','libellé','objet']): col_map[i] = "Désignation"
                elif any(k in ct for k in ['unité','unite','u ']): col_map[i] = "Unité"
                elif any(k in ct for k in ['quantité','quantite','qte']): col_map[i] = "Quantité"
                elif any(k in ct for k in ['prix unitaire','pu ']): col_map[i] = "Prix Unitaire HT"
                elif any(k in ct for k in ['total','montant','prix total']): col_map[i] = "Total HT"
            if len(col_map) < 3: continue
            for row in grid[header_idx+1:]:
                item = {}
                for ci, field in col_map.items():
                    item[field] = str(row[ci]).strip() if ci < len(row) and row[ci] else ""
                if item.get("N° Prix") or item.get("Désignation"): all_items.append(item)
        wb.close()
        return {"document_level":{},"items":all_items,"metadata":{"items_count":len(all_items)}}
    except Exception as e:
        return {"document_level":{},"items":[],"metadata":{"error":str(e)}}

def extract_bp_from_word(file_bytes, ext=".docx"):
    try:
        import docx
        doc = docx.Document(BytesIO(file_bytes))
        all_items = []
        for table in doc.tables:
            grid = [[cell.text.strip() if cell.text else "" for cell in row.cells] for row in table.rows]
            if not grid: continue
            header_idx = None
            for idx, row in enumerate(grid[:25]):
                row_text = " ".join(str(c) for c in row if c).lower()
                if sum(1 for t in ["prix","n°","designation","unité","quantité","total"] if t in row_text) >= 3:
                    header_idx = idx; break
            if header_idx is None: continue
            header = grid[header_idx]
            col_map = {}
            for i, cell in enumerate(header):
                ct = str(cell).lower().strip()
                if any(k in ct for k in ['n°','num','article']): col_map[i] = "N° Prix"
                elif any(k in ct for k in ['designation','description','libellé']): col_map[i] = "Désignation"
                elif any(k in ct for k in ['unité','unite']): col_map[i] = "Unité"
                elif any(k in ct for k in ['quantité','quantite']): col_map[i] = "Quantité"
                elif any(k in ct for k in ['prix unitaire','pu ']): col_map[i] = "Prix Unitaire HT"
                elif any(k in ct for k in ['total','montant']): col_map[i] = "Total HT"
            if len(col_map) < 3: continue
            for row in grid[header_idx+1:]:
                item = {}
                for ci, field in col_map.items():
                    item[field] = str(row[ci]).strip() if ci < len(row) and row[ci] else ""
                if item.get("N° Prix") or item.get("Désignation"): all_items.append(item)
        return {"document_level":{},"items":all_items,"metadata":{"items_count":len(all_items)}}
    except Exception as e:
        return {"document_level":{},"items":[],"metadata":{"error":str(e)}}

def extract_bp_from_pdf(file_bytes, is_scanned=False):
    return {"document_level":{},"items":[],"metadata":{"error":"PDF extraction non supportée en mode simplifié"}}

def extract_bp_fields(file_bytes, file_info):
    ft = file_info.get("type","unknown"); ext = file_info.get("ext","")
    if ft == "excel": result = extract_bp_from_excel(file_bytes, ext)
    elif ft == "word": result = extract_bp_from_word(file_bytes, ext)
    elif ft == "pdf": result = extract_bp_from_pdf(file_bytes)
    else: return {"document_level":{},"items":[],"metadata":{"error":f"Type non supporté: {ft}"}}
    result["extraction_status"] = "success" if result.get("items") else "no_data"
    return result

def process_bp_file(filename, file_bytes):
    ext = Path(filename).suffix.lower()
    ft = "excel" if ext in ['.xlsx','.xlsm','.xls'] else "word" if ext in ['.docx','.doc'] else "pdf" if ext == '.pdf' else None
    if not ft: return {"success":False,"error":f"Type non supporté: {ext}"}
    result = extract_bp_fields(file_bytes, {"type":ft,"ext":ext,"filename":filename})
    if result.get("extraction_status") == "success":
        return {"success":True,"filename":filename,"items_count":len(result.get("items",[])),"bp_result":result}
    return {"success":False,"filename":filename,"error":result.get("metadata",{}).get("error","Échec"),"bp_result":result}

# ═══════════════ MAIN ═══════════════
def process_all_tenders(limit=None, force=False):
    print(f"\n{Back.MAGENTA}{Fore.WHITE}{Style.BRIGHT}{'='*80}")
    print(f"{Back.MAGENTA}{Fore.WHITE}{Style.BRIGHT}  TRAITEMENT BP PAR LOT" + (" ⚡ FORCÉ" if force else ""))
    print(f"{Back.MAGENTA}{Fore.WHITE}{Style.BRIGHT}{'='*80}")
    tenders = get_all_tenders_to_process(force=force)
    if not tenders: print_warning("Aucun AO"); return
    if limit: tenders = tenders[:limit]
    total = len(tenders); print_header(f"🚀 {total} AO"); print_warning("Sauvegarde automatique")
    start = datetime.now()
    for i, tender in enumerate(tenders, 1):
        ref = tender.get('reference'); obj = tender.get('objet','N/A')[:60]
        print(f"\n{Back.CYAN}{Fore.WHITE} AO {i}/{total}: {ref} | {obj}")
        td = get_tender_files(ref)
        if not td or not td.get("files"): global_stats["errors"] += 1; continue
        best, mx = None, 0
        for fd in td["files"]:
            r = process_bp_file(fd["filename"], fd["file_bytes"])
            if r.get("success") and r.get("items_count",0) > mx: mx = r["items_count"]; best = r
        if best:
            bp = best["bp_result"]; doc = bp.get("document_level",{})
            display_bp_summary(doc, mx); print_info(f"Items: {mx}")
            if save_bp_to_supabase(ref, bp, best["filename"]): print_success(f"✅ Sauvegardé ({mx} items)"); global_stats["total_items_extracted"] += mx; global_stats["processed"] += 1
            else: print_error("❌ Sauvegarde échouée"); global_stats["errors"] += 1
        else: global_stats["errors"] += 1
        elapsed = datetime.now() - start; remaining = (elapsed/i)*(total-i) if i > 0 else elapsed
        print_stat(f"Progrès: {i}/{total} | Succès: {global_stats['processed']} | Erreurs: {global_stats['errors']} | Items: {global_stats['total_items_extracted']} | Restant: ~{str(remaining).split('.')[0]}")
    print_header(f"✅ TERMINÉ - {global_stats['processed']} succès, {global_stats['errors']} erreurs, {global_stats['total_items_extracted']} items")

def main():
    parser = argparse.ArgumentParser(description="BP Extractor v3.3")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--all", action="store_true"); group.add_argument("--reference","-r",type=str)
    parser.add_argument("--limit","-l",type=int); parser.add_argument("--yes","-y",action="store_true"); parser.add_argument("--force","-f",action="store_true")
    args = parser.parse_args()
    print(f"\n{Back.BLUE}{Fore.WHITE}{Style.BRIGHT}{'='*80}\n{Back.BLUE}{Fore.WHITE}{Style.BRIGHT}  BP EXTRACTOR v3.3\n{Back.BLUE}{Fore.WHITE}{Style.BRIGHT}{'='*80}")
    if args.all: process_all_tenders(limit=args.limit, force=args.force)
    else:
        td = get_tender_files(args.reference)
        if not td or not td.get("files"): print_error("❌ Pas de fichiers BP"); return
        best, mx = None, 0
        for fd in td["files"]:
            r = process_bp_file(fd["filename"], fd["file_bytes"])
            if r.get("success") and r.get("items_count",0) > mx: mx = r["items_count"]; best = r
        if best:
            bp = best["bp_result"]; display_bp_summary(bp.get("document_level",{}), mx); display_bp_items(bp.get("items",[]))
            if args.yes or input(f"  {Fore.YELLOW}⚠️  Sauvegarder ? (o/N): ").strip().lower() in ['o','oui','y','yes']:
                save_bp_to_supabase(args.reference, bp, best["filename"]); print_success("✅ Sauvegardé")
        else: print_error("❌ Aucun résultat")

if __name__ == "__main__":
    main()
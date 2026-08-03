"""
Excel Extractor — Enhanced for Tender Intelligence
Now extracts Excel Tables, Named Ranges, Hidden rows/cols, AutoFilter, Freeze Panes,
Sheet Protection, and adds heuristic Sheet Type classification.
"""

import io
from pathlib import Path
from typing import Optional
from fastapi import HTTPException

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


def _resolve_fill_color(cell) -> Optional[str]:
    try:
        if cell.fill is None:
            return None
        fg = cell.fill.fgColor
        if fg is None:
            return None
        raw = str(fg.rgb) if fg.rgb is not None else None
        if not raw:
            return None
        raw = raw.strip("'\"")
        if raw in ("00000000", "FF000000", "FFFFFFFF", "00FFFFFF", "none", "None"):
            return None
        if len(raw) == 8:
            return f"#{raw[-6:]}"
        if len(raw) == 6:
            return f"#{raw}"
        return None
    except Exception:
        return None


def _classify_sheet(sheet) -> str:
    """
    Heuristic to classify sheet type based on name and content.
    Returns: Pricing, Technical, Administrative, Inventory, Calculation, Reference, or General.
    """
    name = sheet.title.lower()
    # Sample first 10 rows
    sample_text = ""
    for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
        row_text = " ".join(str(cell) for cell in row if cell)
        if row_text.strip():
            sample_text += row_text + " "

    pricing_kw = ["prix", "bpu", "bp", "dqe", "dp", "quantité", "unitaire", "total", "montant", "devise", "dh", "mad"]
    tech_kw = ["équipement", "matériel", "specification", "technique", "données", "performance", "puissance", "débit", "pression", "tension"]
    admin_kw = ["administratif", "conditions", "clauses", "attestation", "certificat", "qualification", "référence"]
    inv_kw = ["inventaire", "liste", "stock", "quantité", "désignation", "référence"]
    calc_kw = ["calcul", "estimation", "budget", "prévision", "coût"]
    ref_kw = ["référence", "annexe", "document", "index"]

    scores = {
        "Pricing": sum(1 for kw in pricing_kw if kw in sample_text or kw in name),
        "Technical": sum(1 for kw in tech_kw if kw in sample_text or kw in name),
        "Administrative": sum(1 for kw in admin_kw if kw in sample_text or kw in name),
        "Inventory": sum(1 for kw in inv_kw if kw in sample_text or kw in name),
        "Calculation": sum(1 for kw in calc_kw if kw in sample_text or kw in name),
        "Reference": sum(1 for kw in ref_kw if kw in sample_text or kw in name),
    }
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "General"


def extract_excel_structured(file_bytes: bytes, file_name: str = "") -> dict:
    ext = Path(file_name).suffix.lower()
    sheets = []

    if ext in [".xlsx", ".xlsm"]:
        if openpyxl is None:
            raise HTTPException(422, "openpyxl not installed. Run: pip install openpyxl")

        try:
            wb_formulas = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
            wb_values = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            raise HTTPException(422, f"Invalid Excel file: {e}")

        total_rows = 0
        total_formula_cells = 0
        total_merged_cells = 0

        # Extract Named Ranges
        named_ranges = {}
        for name, range_obj in wb_formulas.defined_names.items():
            try:
                for dest in range_obj.destinations:
                    named_ranges[name] = {"sheet": dest[0], "range": str(dest[1])}
            except:
                pass

        for ws_f, ws_v in zip(wb_formulas.worksheets, wb_values.worksheets):
            # Sheet metadata
            sheet_meta = {
                "name": ws_f.title,
                "hidden": ws_f.sheet_state == "hidden",
                "very_hidden": ws_f.sheet_state == "veryHidden",
                "freeze_panes": str(ws_f.freeze_panes) if ws_f.freeze_panes else None,
                "auto_filter": str(ws_f.auto_filter.ref) if ws_f.auto_filter else None,
                "protection": ws_f.protection.enabled if ws_f.protection else False,
                "sheet_type": _classify_sheet(ws_f),
            }

            # Extract Excel Tables
            tables = []
            for table in ws_f.tables.values():
                table_data = {
                    "name": table.name,
                    "display_name": table.displayName,
                    "ref": table.ref,
                    "headers": [],
                    "rows": [],
                }
                try:
                    min_col, min_row, max_col, max_row = table.ref.bounds
                    for col_idx in range(min_col, max_col + 1):
                        cell = ws_v.cell(row=min_row, column=col_idx)
                        table_data["headers"].append(cell.value)
                    for row_idx in range(min_row + 1, max_row + 1):
                        row = []
                        for col_idx in range(min_col, max_col + 1):
                            cell = ws_v.cell(row=row_idx, column=col_idx)
                            row.append(cell.value)
                        table_data["rows"].append(row)
                except:
                    table_data["ref"] = table.ref
                tables.append(table_data)

            # Full grid
            max_row = ws_f.max_row
            max_col = ws_f.max_column
            rows = []
            hidden_rows = set()
            hidden_cols = set()
            for idx in range(1, max_row + 1):
                if ws_f.row_dimensions[idx].hidden:
                    hidden_rows.add(idx)
            for idx in range(1, max_col + 1):
                col_letter = openpyxl.utils.get_column_letter(idx)
                if ws_f.column_dimensions[col_letter].hidden:
                    hidden_cols.add(idx)

            for r in range(1, max_row + 1):
                row = []
                for c in range(1, max_col + 1):
                    cell_f = ws_f.cell(row=r, column=c)
                    cell_v = ws_v.cell(row=r, column=c)
                    is_formula = isinstance(cell_f.value, str) and cell_f.value.startswith("=")
                    if is_formula:
                        total_formula_cells += 1
                    value = cell_v.value if cell_v.value is not None else cell_f.value
                    fill_color = _resolve_fill_color(cell_f)
                    bold = bool(cell_f.font and cell_f.font.bold)
                    italic = bool(cell_f.font and cell_f.font.italic)
                    align = cell_f.alignment.horizontal if cell_f.alignment else None
                    number_format = cell_f.number_format

                    row.append({
                        "row": r,
                        "col": c,
                        "value": value,
                        "formula": cell_f.value if is_formula else None,
                        "bold": bold,
                        "italic": italic,
                        "fill": fill_color,
                        "align": align,
                        "number_format": number_format,
                        "hidden_row": r in hidden_rows,
                        "hidden_col": c in hidden_cols,
                    })
                rows.append(row)

            merges = [str(m) for m in ws_f.merged_cells.ranges]
            total_merged_cells += len(merges)
            total_rows += max_row

            sheets.append({
                "name": ws_f.title,
                "rows": rows,
                "row_count": max_row,
                "col_count": max_col,
                "merged_cells": merges,
                "tables": tables,
                "hidden_rows": list(hidden_rows),
                "hidden_cols": list(hidden_cols),
                "sheet_metadata": sheet_meta,
            })

        stats = {
            "sheet_count": len(sheets),
            "row_count": total_rows,
            "merged_cells": total_merged_cells,
            "formula_cells": total_formula_cells,
            "named_ranges": named_ranges,
        }
        return {"sheets": sheets, "stats": stats, "type": "xlsx"}

    elif ext == ".xls":
        if xlrd is None:
            raise HTTPException(422, "xlrd not installed. Run: pip install xlrd")
        try:
            wb = xlrd.open_workbook(file_contents=file_bytes)
        except Exception as e:
            raise HTTPException(422, f"Invalid XLS file: {e}")

        total_rows = 0
        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            rows = []
            for row_idx in range(sheet.nrows):
                row = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    value = cell.value
                    if cell.ctype == xlrd.sheet.XL_CELL_DATE:
                        try:
                            date_tuple = xlrd.xldate_as_tuple(value, wb.datemode)
                            if date_tuple[0] > 0:
                                value = f"{date_tuple[0]:04d}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                            else:
                                value = ""
                        except:
                            value = str(value)
                    elif cell.ctype == xlrd.sheet.XL_CELL_BOOLEAN:
                        value = "TRUE" if value else "FALSE"
                    elif cell.ctype == xlrd.sheet.XL_CELL_EMPTY:
                        value = ""
                    elif cell.ctype == xlrd.sheet.XL_CELL_NUMBER:
                        if isinstance(value, float) and value.is_integer():
                            value = int(value)
                    row.append({
                        "row": row_idx + 1,
                        "col": col_idx + 1,
                        "value": value,
                        "formula": None,
                        "bold": False,
                        "italic": False,
                        "fill": None,
                        "align": None,
                        "number_format": None,
                        "hidden_row": False,
                        "hidden_col": False,
                    })
                rows.append(row)
            total_rows += sheet.nrows
            sheets.append({
                "name": sheet.name,
                "rows": rows,
                "row_count": sheet.nrows,
                "col_count": sheet.ncols,
                "merged_cells": [],
                "tables": [],
                "hidden_rows": [],
                "hidden_cols": [],
                "sheet_metadata": {"hidden": False, "very_hidden": False, "freeze_panes": None,
                                    "auto_filter": None, "protection": False, "sheet_type": "General"},
            })

        stats = {
            "sheet_count": len(sheets),
            "row_count": total_rows,
            "merged_cells": 0,
            "formula_cells": 0,
            "named_ranges": {},
        }
        return {"sheets": sheets, "stats": stats, "type": "xls"}

    else:
        raise HTTPException(400, "Unsupported Excel format")